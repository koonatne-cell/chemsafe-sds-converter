# -*- coding: utf-8 -*-
"""
fill_excel.py - วางข้อมูลลงบน Template.xlsx (ไฟล์ต้นแบบจริงของบริษัท) แล้วเซฟไฟล์ .xlsx ใหม่

ต่างจาก fill_template.py (PDF) ตรงที่ไม่ต้องเดาพิกัด (x, y) เอง หรือคอยลดฟอนต์กันข้อความล้น
(ตั้ง wrap_text=True ให้ทุกช่องตอนเขียนค่า กันข้อความบรรทัดเดียวล้นไปทับ cell ข้างๆ ที่มีป้ายอื่นอยู่)
ผู้ใช้เปิดไฟล์ที่ได้ด้วย Microsoft Excel บนเครื่องตัวเอง แล้วปริ้นเองอีกที

หมายเหตุ: แถวในเทมเพลตถูกล็อกความสูงไว้ตายตัว (customHeight) ถ้าข้อความยาวเกินจะถูกตัดบรรทัด
แต่อาจไม่พอดีกับความสูงแถว ผู้ใช้ปรับความสูงแถวเองได้ใน Excel (Home > Format > AutoFit Row Height)

หมายเหตุ: ตัวเลขดัชนี NFPA (สุขภาพ/ไวไฟ/ปฏิกิริยา) ในไฟล์ต้นแบบจริงอยู่ใน "shape รูปเพชร"
ไม่ใช่ cell ธรรมดา (openpyxl แก้ข้อความใน shape ไม่ได้โดยตรง) จึงเว้นว่างไว้ให้ผู้ใช้กรอกเองใน Excel
"""
import os
import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

# PROJECT_ROOT ต้องขึ้นไปอีก 1 ชั้นจากไฟล์นี้ เพราะย้ายเข้ามาอยู่ใน pdf_gen/ แล้ว (ไม่ได้ใช้ตัวแปรนี้
# ที่ไหนตอนนี้ แต่แก้ไว้เผื่ออนาคต - ดู pdf_gen/fill_template.py สำหรับเหตุผลเดียวกัน)
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SHEET_NAME = "1.00.05"

# แผนที่ field key -> ตำแหน่ง cell จริงในเทมเพลต (วัดจากไฟล์ต้นแบบของบริษัท)
CELL_MAP = {
    "display_name": "H6",   # แถบเหลือง: ชื่อสารเคมี (ตัวใหญ่)
    "signal_word":  "I8",   # แถบแดง: signal word (เช่น "อันตราย" / "สารระคายเคือง")
    "trade_name":   "A11",  # "ชื่อสารเคมี :<ค่า>" (พิมพ์รวมกับป้ายในเซลล์เดียว ตามต้นฉบับ)
    "formula":      "U11",
    "un":           "AB11",
    "cas":          "AG11",
    "state":        "F12",
    "color":        "L12",
    "odor":         "P12",
    "boiling":      "U12",
    "ph":           "AB12",
    "flash":        "AG12",
    "usage":        "F13",
    "reactivity":   "F14",
    "fire":         "F17",
    "hz_eye":       "F19",
    "hz_oral":      "F20",
    "hz_skin":      "F21",
    "hz_inhale":    "F22",
    "fa_eye":       "F24",
    "fa_oral":      "F25",
    "fa_skin":      "F26",
    "fa_inhale":    "F27",
    "spill":        "F28",
    "disposal":     "F31",
    "storage":      "F34",
    "prepared_name":     "I49",
    "prepared_position": "I50",
    "approved_name":     "AG49",
    "approved_position": "AD50",
}

# เซลล์ที่พิมพ์ป้ายรวมกับค่าในเซลล์เดียวกัน (ต้องเติมป้ายกลับเข้าไปเองตอนวางค่า)
LABEL_PREFIX = {
    "trade_name": "ชื่อสารเคมี :",
}

# ตำแหน่งกล่องรูปภาพ (แถว 40-45 คอลัมน์ A-U สำหรับรูปสลาก, X-AO สำหรับรูปภาชนะ)
# ใช้ TwoCellAnchor แบบยืดรูปให้เต็มกรอบ (ไม่รักษาสัดส่วนเป๊ะ เพราะความกว้างคอลัมน์ของไฟล์นี้
# วัดเป็นพิกเซลได้ไม่แม่นยำ ผู้ใช้ปรับขนาดรูปเองใน Excel ได้ถ้าต้องการ)
IMAGE_ANCHORS = {
    "label_image":     (0, 39, 21, 45),   # A40:U45 (คอลัมน์/แถว นับจาก 0)
    "container_image": (23, 39, 41, 45),  # X40:AO45
}


def _set_cell(ws, key, value):
    if value is None:
        return
    value = str(value).strip()
    if not value or value == "-":
        return
    addr = CELL_MAP[key]
    prefix = LABEL_PREFIX.get(key, "")
    cell = ws[addr]
    cell.value = prefix + value
    # เปิด wrap_text เสมอ กันข้อความยาวบรรทัดเดียวล้นไปทับ cell ข้างๆ ที่มีป้าย/ข้อมูลอื่นอยู่
    # (คงการจัดวางแนวนอน/แนวตั้งเดิมของเทมเพลตไว้ เปลี่ยนแค่ wrap_text)
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal=existing.horizontal, vertical=existing.vertical, wrap_text=True,
    )


def _add_image(ws, image_path, anchor_key):
    if not image_path or not os.path.exists(image_path):
        return
    col_start, row_start, col_end, row_end = IMAGE_ANCHORS[anchor_key]
    img = XLImage(image_path)
    marker_from = AnchorMarker(col=col_start, colOff=0, row=row_start, rowOff=0)
    marker_to = AnchorMarker(col=col_end, colOff=0, row=row_end, rowOff=0)
    img.anchor = TwoCellAnchor(editAs="oneCell", _from=marker_from, to=marker_to)
    ws.add_image(img)


def fill_from_data(data, template_path, out_path, label_image_path=None, container_image_path=None):
    """รับ data dict ที่แก้ไขแล้วจาก UI + รูป (ถ้ามี) วางลงเทมเพลตแล้วเซฟไฟล์ .xlsx ใหม่"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_NAME]

    for key in CELL_MAP:
        _set_cell(ws, key, data.get(key))

    # วันที่ออกเอกสาร ใช้วันที่สร้างจริง ไม่ใช้ค่าค้างจากไฟล์ต้นแบบ
    ws["Y2"] = datetime.date.today()

    _add_image(ws, label_image_path, "label_image")
    _add_image(ws, container_image_path, "container_image")

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    # รันจากรากโปรเจกต์ด้วย: python -m pdf_gen.fill_excel (import แบบ package ต้องใช้ -m)
    from core.parser import parse_sds
    data = parse_sds("getpdf_sample.pdf") if os.path.exists("getpdf_sample.pdf") else {}
    fill_from_data(data, "assets/Template.xlsx", "data/generated/test_output.xlsx")
    print("saved test_output.xlsx")
